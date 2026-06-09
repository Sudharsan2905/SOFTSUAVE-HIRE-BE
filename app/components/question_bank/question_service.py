from typing import Any, cast

from bson import ObjectId
from motor.motor_asyncio import AsyncIOMotorDatabase

from app.common.constants.app_constants import QuestionType
from app.common.constants.messages import ErrorMessages
from app.common.exceptions import ConflictException, NotFoundException, ValidationException
from app.common.response_models.question_responses import (
    BulkOperationResponse,
    CategoryResponse,
    QuestionResponse,
)
from app.common.utils import (
    build_pagination_meta,
    list_paginated,
    paginate_query,
    safe_regex,
    serialize_doc,
    serialize_docs,
    utcnow,
)
from app.core.logging import logger

_REGEX = "$regex"
_OPTIONS = "$options"


def _validate_question_options(data: dict) -> None:
    """Enforce MCQ/essay option rules before insert or update."""
    q_type = data.get("question_type", "")
    options = data.get("options") or []

    if q_type in [QuestionType.MCQ_SINGLE, QuestionType.MCQ_MULTI]:
        if not options:
            raise ValidationException("MCQ questions must have at least one option")
        if not any(
            o.get("is_correct") if isinstance(o, dict) else getattr(o, "is_correct", False)
            for o in options
        ):
            raise ValidationException("MCQ questions must have at least one correct option marked")

    if q_type == QuestionType.ESSAY and options:
        raise ValidationException("Essay questions must not have options")


async def create_category(db: AsyncIOMotorDatabase, data: dict, user_id: str) -> CategoryResponse:
    """Create a new question category (case-insensitive duplicate check).

    Raises:
        ConflictException: If a category with the same name already exists.
    """
    if await db.question_categories.find_one(
        {"name": {_REGEX: f"^{safe_regex(data['name'])}$", _OPTIONS: "i"}}
    ):
        raise ConflictException(f"Category '{data['name']}' already exists")

    now = utcnow()
    doc = {
        "name": data["name"],
        "description": data.get("description", ""),
        "created_by": ObjectId(user_id),
        "question_count": 0,
        "created_at": now,
        "updated_at": now,
    }
    result = await db.question_categories.insert_one(doc)
    doc["_id"] = result.inserted_id
    return CategoryResponse.model_validate(serialize_doc(doc))


async def get_categories(
    db: AsyncIOMotorDatabase,
    search: str | None,
    sort_by: str,
    sort_order: str,
    page: int,
    page_size: int,
) -> dict:
    """Return a paginated list of question categories, optionally filtered by name."""
    skip, limit = paginate_query(page, page_size)
    query = {}
    if search:
        query["name"] = {_REGEX: safe_regex(search), _OPTIONS: "i"}

    sort_dir = 1 if sort_order == "asc" else -1
    total, docs = await list_paginated(
        db.question_categories,
        query,
        sort_by,
        sort_dir,
        skip,
        limit,
        ["name", "created_at", "updated_at", "question_count"],
    )
    return {
        "categories": serialize_docs(docs),
        "pagination": build_pagination_meta(total, page, page_size),
    }


async def update_category(
    db: AsyncIOMotorDatabase, category_id: str, data: dict
) -> CategoryResponse:
    """Update a category's name or description.

    Raises:
        NotFoundException: If the category does not exist.
    """
    if not await db.question_categories.find_one({"_id": ObjectId(category_id)}):
        raise NotFoundException(ErrorMessages.CATEGORY_NOT_FOUND)
    update = {k: v for k, v in data.items() if v is not None}
    update["updated_at"] = utcnow()
    await db.question_categories.update_one({"_id": ObjectId(category_id)}, {"$set": update})
    updated = await db.question_categories.find_one({"_id": ObjectId(category_id)})
    return CategoryResponse.model_validate(serialize_doc(updated))


async def delete_category(db: AsyncIOMotorDatabase, category_id: str) -> None:
    """Delete a category and cascade-delete all its questions.

    Raises:
        NotFoundException: If the category does not exist.
    """
    if not await db.question_categories.find_one({"_id": ObjectId(category_id)}):
        raise NotFoundException(ErrorMessages.CATEGORY_NOT_FOUND)
    await db.question_categories.delete_one({"_id": ObjectId(category_id)})
    await db.questions.delete_many({"category_id": ObjectId(category_id)})


async def get_questions(
    db: AsyncIOMotorDatabase,
    category_id: str,
    search: str | None,
    complexity: str | None,
    question_type: str | None,
    sort_by: str,
    sort_order: str,
    page: int,
    page_size: int,
) -> dict:
    """Return a paginated list of questions in a category, with optional filters."""
    skip, limit = paginate_query(page, page_size)
    query: dict = {"category_id": ObjectId(category_id)}
    if search:
        query["question_text"] = {_REGEX: safe_regex(search), _OPTIONS: "i"}
    if complexity:
        query["complexity"] = complexity
    if question_type:
        query["question_type"] = question_type

    sort_dir = 1 if sort_order == "asc" else -1
    total, docs = await list_paginated(
        db.questions,
        query,
        sort_by,
        sort_dir,
        skip,
        limit,
        ["created_at", "updated_at", "complexity"],
    )
    return {
        "questions": serialize_docs(docs),
        "pagination": build_pagination_meta(total, page, page_size),
    }


async def create_question(
    db: AsyncIOMotorDatabase, category_id: str, data: dict, user_id: str
) -> QuestionResponse:
    """Create a single question in a category, enforcing MCQ/essay option rules.

    Raises:
        NotFoundException: If the category does not exist.
        ValidationException: If MCQ has no correct option, or essay has options.
    """
    if not await db.question_categories.find_one({"_id": ObjectId(category_id)}):
        raise NotFoundException(ErrorMessages.CATEGORY_NOT_FOUND)

    _validate_question_options(data)
    now = utcnow()
    options = data.get("options") or []
    if isinstance(options, list):
        options = [o.model_dump() if hasattr(o, "model_dump") else o for o in options]

    doc = {
        "category_id": ObjectId(category_id),
        "question_text": data["question_text"],
        "question_type": data["question_type"],
        "complexity": data["complexity"],
        "options": options,
        "correct_answer": data.get("correct_answer"),
        "created_by": ObjectId(user_id),
        "created_at": now,
        "updated_at": now,
    }
    result = await db.questions.insert_one(doc)
    doc["_id"] = result.inserted_id
    await db.question_categories.update_one(
        {"_id": ObjectId(category_id)}, {"$inc": {"question_count": 1}}
    )
    return QuestionResponse.model_validate(serialize_doc(doc))


async def bulk_create_questions(
    db: AsyncIOMotorDatabase, category_id: str, questions: list, user_id: str
) -> BulkOperationResponse:
    """Insert multiple questions at once and increment the category's question_count.

    Returns:
        Dict with key 'created' containing the number of inserted questions.

    Raises:
        NotFoundException: If the category does not exist.
    """
    if not await db.question_categories.find_one({"_id": ObjectId(category_id)}):
        raise NotFoundException(ErrorMessages.CATEGORY_NOT_FOUND)

    now = utcnow()
    docs = []
    for q in questions:
        options = q.get("options") or []
        if isinstance(options, list):
            options = [o.model_dump() if hasattr(o, "model_dump") else o for o in options]
        docs.append(
            {
                "category_id": ObjectId(category_id),
                "question_text": q["question_text"],
                "question_type": q["question_type"],
                "complexity": q["complexity"],
                "options": options,
                "correct_answer": q.get("correct_answer"),
                "created_by": ObjectId(user_id),
                "created_at": now,
                "updated_at": now,
            }
        )

    if docs:
        result = await db.questions.insert_many(docs)
        await db.question_categories.update_one(
            {"_id": ObjectId(category_id)}, {"$inc": {"question_count": len(docs)}}
        )
        return BulkOperationResponse(created=len(result.inserted_ids))
    return BulkOperationResponse(created=0)


async def update_question(
    db: AsyncIOMotorDatabase, question_id: str, data: dict
) -> QuestionResponse:
    """Update a question's fields, re-validating MCQ/essay option rules.

    Raises:
        NotFoundException: If the question does not exist.
        ValidationException: If updated options violate MCQ/essay rules.
    """
    if not await db.questions.find_one({"_id": ObjectId(question_id)}):
        raise NotFoundException("Question not found")
    _validate_question_options(data)
    update = {k: v for k, v in data.items() if v is not None}
    if "options" in update and isinstance(update["options"], list):
        update["options"] = [
            o.model_dump() if hasattr(o, "model_dump") else o for o in update["options"]
        ]
    update["updated_at"] = utcnow()
    await db.questions.update_one({"_id": ObjectId(question_id)}, {"$set": update})
    updated = await db.questions.find_one({"_id": ObjectId(question_id)})
    return QuestionResponse.model_validate(serialize_doc(updated))


async def delete_question(db: AsyncIOMotorDatabase, question_id: str) -> None:
    """Delete a question and decrement its category's question_count.

    Raises:
        NotFoundException: If the question does not exist.
    """
    q = await db.questions.find_one({"_id": ObjectId(question_id)})
    if not q:
        raise NotFoundException(ErrorMessages.QUESTION_NOT_FOUND)
    await db.questions.delete_one({"_id": ObjectId(question_id)})
    await db.question_categories.update_one(
        {"_id": q["category_id"]}, {"$inc": {"question_count": -1}}
    )


async def ai_generate_questions(
    db: AsyncIOMotorDatabase,
    category_id: str,
    topic: str,
    count: int,
    complexity: str,
    question_type: str,
    user_id: str,
) -> BulkOperationResponse:
    """Generate questions via OpenAI GPT-4o and bulk-insert them into the category.

    Returns:
        Dict with 'created' count, or 'error' key if generation fails.

    Raises:
        NotFoundException: If the category does not exist (raised by bulk_create_questions).
    """
    try:
        import json
        import re

        import openai

        from app.core.config import settings

        client: Any = openai.OpenAI(api_key=settings.OPENAI_API_KEY)
        type_instructions = {
            "mcq_single": (
                "MCQ with exactly one correct answer. "
                "Provide 4 options with field 'is_correct': true/false."
            ),
            "mcq_multi": (
                "MCQ with one or more correct answers. "
                "Provide 4 options with field 'is_correct': true/false."
            ),
            "essay": "Open-ended essay question. Provide a 'correct_answer' with a model answer.",
        }

        prompt = f"""
Generate {count} {complexity}-difficulty technical interview questions on the topic: "{topic}".
Question type: {question_type} - {type_instructions.get(question_type, "")}

Return ONLY a valid JSON array. Each object must have:
- question_text (string) — if the question contains code, wrap it in a markdown fenced code block
    with the appropriate language tag (e.g. ```python\\n...\\n```)
- options (array of {{id, text, is_correct}}) — only for MCQ types
- correct_answer (string) — only for essay type

No extra explanation outside the JSON array. Return only the JSON array."""

        message = client.chat.completions.create(
            model="gpt-4o",
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )
        content = message.choices[0].message.content
        if not content:
            return BulkOperationResponse(created=0, error="AI generation failed or unavailable")
        json_match = re.search(r"\[.*\]", content, re.DOTALL)
        if json_match:
            questions_data = json.loads(json_match.group())
            enriched = [
                {**q, "question_type": question_type, "complexity": complexity}
                for q in questions_data
            ]
            result = await bulk_create_questions(db, category_id, enriched, user_id)
            logger.info(
                f"AI generated {result.created} questions for category_id={category_id} "
                f"topic='{topic}' type={question_type}"
            )
            return result
    except Exception:
        logger.exception(
            f"AI question generation failed for category_id={category_id} topic='{topic}'"
        )

    return BulkOperationResponse(created=0, error="AI generation failed or unavailable")


def _find_column(headers_lower: list, name: str) -> str | None:
    """Case-insensitive column lookup; returns original-case header or None."""
    for h in headers_lower:
        if h and cast(str, h).lower() == name.lower():
            return cast(str, h)
    return None


def _split_outside_brackets(text: str) -> list[str]:
    """Split on commas that are not inside any bracket pair ( [ {."""
    result = []
    current: list[str] = []
    stack: list[str] = []
    pairs = {")": "(", "]": "[", "}": "{"}
    for ch in text:
        if ch in "([{":
            stack.append(ch)
        elif ch in ")]}":
            if stack and stack[-1] == pairs[ch]:
                stack.pop()
        if ch == "," and not stack:
            result.append("".join(current).strip())
            current = []
        else:
            current.append(ch)
    if current:
        result.append("".join(current).strip())
    return result


def _classify_and_build(
    options_raw: object, answer_raw: object
) -> tuple[str, list[dict], str | None]:
    """Return (question_type, options_list, correct_answer_text)."""
    options_str = str(options_raw).strip() if options_raw else ""
    answer_str = str(answer_raw).strip() if answer_raw is not None else ""

    if not options_str:
        # No options → essay
        return "essay", [], answer_str or None

    # Split on commas outside brackets so options like print("a, b") stay intact
    option_texts = [p for p in _split_outside_brackets(options_str) if p]

    # Determine correct indices (1-based numbers in answer)
    correct_indices: set[int] = set()
    for part in answer_str.split(","):
        try:
            correct_indices.add(int(part.strip()))
        except ValueError:
            pass

    question_type = "mcq_multi" if len(correct_indices) > 1 else "mcq_single"

    import uuid

    options = [
        {"id": str(uuid.uuid4()), "text": text, "is_correct": (i + 1) in correct_indices}
        for i, text in enumerate(option_texts)
    ]
    return question_type, options, None


async def process_excel_import(
    db: AsyncIOMotorDatabase,
    category_id: str,
    file_data: bytes,
    user_id: str,
    column_map: dict | None = None,
) -> BulkOperationResponse:
    """Parse an Excel file and bulk-import questions into a category.

    Args:
        db: AsyncIOMotorDatabase instance.
        category_id: Target category string ID.
        file_data: Raw bytes of the uploaded .xlsx file.
        user_id: ID of the admin performing the import.
        column_map: Optional mapping of logical fields to Excel column headers
            (keys: 'question', 'options', 'answer', 'complexity'). Falls back to
            case-insensitive header auto-detection when not provided.

    Returns:
        Dict with 'created' count, or 'error' key on failure.
    """
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_data))
    ws = wb.active
    raw_headers = [cell.value for cell in ws[1]]

    # Use explicit column_map first, fall back to case-insensitive auto-detect
    col_map = column_map or {}
    col_question = col_map.get("question") or _find_column(raw_headers, "question")
    col_options = col_map.get("options") or _find_column(raw_headers, "options")
    col_answer = col_map.get("answer") or _find_column(raw_headers, "answer")
    col_complexity = col_map.get("complexity") or _find_column(raw_headers, "complexity")

    if not col_question:
        return BulkOperationResponse(created=0, error="Missing 'Question' column")

    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(raw_headers, row, strict=False))

        question_text = row_dict.get(col_question)
        if not question_text:
            continue

        options_raw = row_dict.get(col_options) if col_options else None
        answer_raw = row_dict.get(col_answer) if col_answer else None
        complexity_raw = str(row_dict.get(col_complexity, "")).lower() if col_complexity else ""
        complexity = complexity_raw if complexity_raw in ("low", "medium", "high") else "medium"

        question_type, options, correct_answer = _classify_and_build(options_raw, answer_raw)

        questions.append(
            {
                "question_text": str(question_text).strip(),
                "question_type": question_type,
                "complexity": complexity,
                "options": options,
                "correct_answer": correct_answer,
            }
        )

    return await bulk_create_questions(db, category_id, questions, user_id)
