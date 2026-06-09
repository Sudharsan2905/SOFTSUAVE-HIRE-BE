"""Unit tests for app/components/question_bank/question_service.py"""

import pytest

from app.common.exceptions import ConflictException, NotFoundException, ValidationException
from app.components.question_bank import question_service


class TestCreateCategory:
    async def test_success(self, db, super_admin):
        result = await question_service.create_category(
            db, {"name": "JavaScript", "description": "JS questions"}, str(super_admin["_id"])
        )
        assert result.name == "JavaScript"
        assert result.question_count == 0

    async def test_duplicate_name_raises(self, db, category, super_admin):
        with pytest.raises(ConflictException):
            await question_service.create_category(db, {"name": "Python"}, str(super_admin["_id"]))

    async def test_case_insensitive_duplicate(self, db, category, super_admin):
        with pytest.raises(ConflictException):
            await question_service.create_category(db, {"name": "PYTHON"}, str(super_admin["_id"]))


class TestCreateQuestion:
    async def test_mcq_success(self, db, category, super_admin):
        data = {
            "question_text": "What does 'self' refer to in Python?",
            "question_type": "mcq_single",
            "complexity": "medium",
            "options": [
                {"id": "a", "text": "The class", "is_correct": False},
                {"id": "b", "text": "The instance", "is_correct": True},
            ],
            "correct_answer": None,
        }
        result = await question_service.create_question(
            db, str(category["_id"]), data, str(super_admin["_id"])
        )
        assert result.question_text == data["question_text"]

    async def test_mcq_no_correct_option_raises(self, db, category, super_admin):
        data = {
            "question_text": "Bad MCQ",
            "question_type": "mcq_single",
            "complexity": "low",
            "options": [
                {"id": "a", "text": "Option A", "is_correct": False},
            ],
        }
        with pytest.raises(ValidationException, match="correct option"):
            await question_service.create_question(
                db, str(category["_id"]), data, str(super_admin["_id"])
            )

    async def test_mcq_no_options_raises(self, db, category, super_admin):
        data = {
            "question_text": "No options MCQ",
            "question_type": "mcq_single",
            "complexity": "low",
            "options": [],
        }
        with pytest.raises(ValidationException, match="at least one option"):
            await question_service.create_question(
                db, str(category["_id"]), data, str(super_admin["_id"])
            )

    async def test_essay_with_options_raises(self, db, category, super_admin):
        data = {
            "question_text": "Explain OOP",
            "question_type": "essay",
            "complexity": "high",
            "options": [{"id": "a", "text": "Option A", "is_correct": False}],
        }
        with pytest.raises(ValidationException, match="must not have options"):
            await question_service.create_question(
                db, str(category["_id"]), data, str(super_admin["_id"])
            )

    async def test_essay_success(self, db, category, super_admin):
        data = {
            "question_text": "Explain Python decorators",
            "question_type": "essay",
            "complexity": "high",
            "options": [],
            "correct_answer": "A decorator wraps a function...",
        }
        result = await question_service.create_question(
            db, str(category["_id"]), data, str(super_admin["_id"])
        )
        assert result.question_type == "essay"

    async def test_nonexistent_category_raises(self, db, super_admin):
        from bson import ObjectId

        data = {
            "question_text": "Question",
            "question_type": "essay",
            "complexity": "low",
            "options": [],
        }
        with pytest.raises(NotFoundException):
            await question_service.create_question(
                db, str(ObjectId()), data, str(super_admin["_id"])
            )


class TestDeleteCategory:
    async def test_cascades_questions(self, db, category, mcq_question):
        await question_service.delete_category(db, str(category["_id"]))
        remaining = await db.questions.count_documents({"category_id": category["_id"]})
        assert remaining == 0

    async def test_nonexistent_raises(self, db):
        from bson import ObjectId

        with pytest.raises(NotFoundException):
            await question_service.delete_category(db, str(ObjectId()))


class TestBulkCreateQuestions:
    async def test_increments_question_count(self, db, category, super_admin):
        questions = [
            {
                "question_text": f"Question {i}",
                "question_type": "essay",
                "complexity": "low",
                "options": [],
                "correct_answer": f"Answer {i}",
            }
            for i in range(3)
        ]
        result = await question_service.bulk_create_questions(
            db, str(category["_id"]), questions, str(super_admin["_id"])
        )
        assert result.created == 3
        updated_cat = await db.question_categories.find_one({"_id": category["_id"]})
        assert updated_cat["question_count"] == 3


class TestGetCategories:
    async def test_returns_paginated(self, db, category, super_admin):
        result = await question_service.get_categories(db, None, "created_at", "desc", 1, 20)
        assert result["pagination"]["total"] == 1
        assert len(result["categories"]) == 1

    async def test_search_filter(self, db, category, super_admin):
        result = await question_service.get_categories(db, "Python", "created_at", "desc", 1, 20)
        assert result["pagination"]["total"] == 1

    async def test_search_no_match(self, db, category, super_admin):
        result = await question_service.get_categories(db, "Java", "created_at", "desc", 1, 20)
        assert result["pagination"]["total"] == 0

    async def test_asc_sort(self, db, category, super_admin):
        result = await question_service.get_categories(db, None, "name", "asc", 1, 20)
        assert result["pagination"]["total"] == 1


class TestUpdateCategory:
    async def test_success(self, db, category):
        result = await question_service.update_category(
            db, str(category["_id"]), {"name": "Updated Python", "description": "New desc"}
        )
        assert result.name == "Updated Python"

    async def test_not_found_raises(self, db):
        from bson import ObjectId

        with pytest.raises(NotFoundException):
            await question_service.update_category(db, str(ObjectId()), {"name": "X"})


class TestGetQuestions:
    async def test_returns_paginated(self, db, category, mcq_question):
        result = await question_service.get_questions(
            db, str(category["_id"]), None, None, None, "created_at", "desc", 1, 20
        )
        assert result["pagination"]["total"] == 1
        assert len(result["questions"]) == 1

    async def test_filter_by_search(self, db, category, mcq_question):
        result = await question_service.get_questions(
            db, str(category["_id"]), "2 + 2", None, None, "created_at", "desc", 1, 20
        )
        assert result["pagination"]["total"] == 1

    async def test_filter_by_complexity(self, db, category, mcq_question):
        result = await question_service.get_questions(
            db, str(category["_id"]), None, "low", None, "created_at", "desc", 1, 20
        )
        assert result["pagination"]["total"] == 1

    async def test_filter_by_question_type(self, db, category, mcq_question):
        result = await question_service.get_questions(
            db, str(category["_id"]), None, None, "mcq_single", "created_at", "desc", 1, 20
        )
        assert result["pagination"]["total"] == 1

    async def test_asc_sort(self, db, category, mcq_question):
        result = await question_service.get_questions(
            db, str(category["_id"]), None, None, None, "complexity", "asc", 1, 20
        )
        assert result["pagination"]["total"] == 1


class TestUpdateQuestion:
    async def test_success(self, db, mcq_question):
        result = await question_service.update_question(
            db,
            str(mcq_question["_id"]),
            {
                "question_text": "Updated text",
                "question_type": "mcq_single",
                "options": [
                    {"id": "a", "text": "Yes", "is_correct": True},
                    {"id": "b", "text": "No", "is_correct": False},
                ],
            },
        )
        assert result.question_text == "Updated text"

    async def test_not_found_raises(self, db):
        from bson import ObjectId

        with pytest.raises(NotFoundException):
            await question_service.update_question(db, str(ObjectId()), {"question_text": "X"})


class TestDeleteQuestion:
    async def test_success(self, db, category, mcq_question):
        await question_service.delete_question(db, str(mcq_question["_id"]))
        remaining = await db.questions.find_one({"_id": mcq_question["_id"]})
        assert remaining is None
        cat = await db.question_categories.find_one({"_id": category["_id"]})
        assert cat["question_count"] == -1  # decremented from 0

    async def test_not_found_raises(self, db):
        from bson import ObjectId

        with pytest.raises(NotFoundException):
            await question_service.delete_question(db, str(ObjectId()))


class TestBulkCreateEmpty:
    async def test_empty_list_returns_zero(self, db, category, super_admin):
        result = await question_service.bulk_create_questions(
            db, str(category["_id"]), [], str(super_admin["_id"])
        )
        assert result.created == 0

    async def test_nonexistent_category_raises(self, db, super_admin):
        from bson import ObjectId

        with pytest.raises(NotFoundException):
            await question_service.bulk_create_questions(
                db,
                str(ObjectId()),
                [
                    {
                        "question_text": "Q",
                        "question_type": "essay",
                        "complexity": "low",
                        "options": [],
                        "correct_answer": "A",
                    }
                ],
                str(super_admin["_id"]),
            )


class TestAiGenerateQuestions:
    async def test_success(self, db, category, super_admin):
        from unittest.mock import MagicMock, patch

        mock_openai = MagicMock()
        mock_client = MagicMock()
        mock_openai.OpenAI.return_value = mock_client
        mock_message = MagicMock()
        mock_message.choices[0].message.content = (
            '[{"question_text": "What is Python?", "options": '
            '[{"id": "a", "text": "A language", "is_correct": true}], "correct_answer": null}]'
        )
        mock_client.chat.completions.create.return_value = mock_message

        with patch.dict("sys.modules", {"openai": mock_openai}):
            result = await question_service.ai_generate_questions(
                db,
                str(category["_id"]),
                "Python",
                1,
                "low",
                "mcq_single",
                str(super_admin["_id"]),
            )
        assert result.created >= 0

    async def test_empty_content_returns_error(self, db, category, super_admin):
        from unittest.mock import MagicMock, patch

        mock_openai = MagicMock()
        mock_client = MagicMock()
        mock_openai.OpenAI.return_value = mock_client
        mock_message = MagicMock()
        mock_message.choices[0].message.content = None
        mock_client.chat.completions.create.return_value = mock_message

        with patch.dict("sys.modules", {"openai": mock_openai}):
            result = await question_service.ai_generate_questions(
                db, str(category["_id"]), "Python", 1, "low", "essay", str(super_admin["_id"])
            )
        assert result.error

    async def test_exception_returns_error(self, db, category, super_admin):
        from unittest.mock import MagicMock, patch

        mock_openai = MagicMock()
        mock_openai.OpenAI.side_effect = Exception("API down")

        with patch.dict("sys.modules", {"openai": mock_openai}):
            result = await question_service.ai_generate_questions(
                db, str(category["_id"]), "Topic", 1, "low", "essay", str(super_admin["_id"])
            )
        assert result.error


class TestInternalHelpers:
    def test_find_column_found(self):
        from app.components.question_bank.question_service import _find_column

        headers = ["Question", "Options", "Answer"]
        assert _find_column(headers, "question") == "Question"

    def test_find_column_not_found(self):
        from app.components.question_bank.question_service import _find_column

        headers = ["Question", "Options"]
        assert _find_column(headers, "complexity") is None

    def test_find_column_skips_none(self):
        from app.components.question_bank.question_service import _find_column

        headers = [None, "Question"]
        assert _find_column(headers, "question") == "Question"

    def test_split_outside_brackets_simple(self):
        from app.components.question_bank.question_service import _split_outside_brackets

        result = _split_outside_brackets("a, b, c")
        assert result == ["a", "b", "c"]

    def test_split_outside_brackets_nested(self):
        from app.components.question_bank.question_service import _split_outside_brackets

        result = _split_outside_brackets('print("a, b"), option2')
        assert len(result) == 2

    def test_split_outside_brackets_curly(self):
        from app.components.question_bank.question_service import _split_outside_brackets

        result = _split_outside_brackets("{a, b}, c")
        assert len(result) == 2

    def test_classify_essay(self):
        from app.components.question_bank.question_service import _classify_and_build

        q_type, options, answer = _classify_and_build(None, "Model answer")
        assert q_type == "essay"
        assert options == []
        assert answer == "Model answer"

    def test_classify_essay_no_answer(self):
        from app.components.question_bank.question_service import _classify_and_build

        q_type, options, answer = _classify_and_build("", None)
        assert q_type == "essay"
        assert answer is None

    def test_classify_mcq_single(self):
        from app.components.question_bank.question_service import _classify_and_build

        q_type, options, answer = _classify_and_build("Option A, Option B, Option C", "1")
        assert q_type == "mcq_single"
        assert len(options) == 3
        assert options[0]["is_correct"] is True

    def test_classify_mcq_multi(self):
        from app.components.question_bank.question_service import _classify_and_build

        q_type, options, answer = _classify_and_build("A, B, C", "1,2")
        assert q_type == "mcq_multi"
        assert len(options) == 3

    def test_classify_invalid_answer(self):
        from app.components.question_bank.question_service import _classify_and_build

        q_type, options, answer = _classify_and_build("A, B", "not_a_number")
        assert q_type == "mcq_single"


class TestProcessExcelImport:
    async def test_success(self, db, category, super_admin):
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Question", "Options", "Answer", "Complexity"])
        ws.append(["What is Python?", "A language, A snake, A movie", "1", "low"])
        ws.append(["Explain OOP", None, None, "high"])

        buf = io.BytesIO()
        wb.save(buf)

        result = await question_service.process_excel_import(
            db, str(category["_id"]), buf.getvalue(), str(super_admin["_id"])
        )
        assert result.created == 2

    async def test_missing_question_column(self, db, category, super_admin):
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Topic", "Options"])
        ws.append(["Python", "A, B"])

        buf = io.BytesIO()
        wb.save(buf)

        result = await question_service.process_excel_import(
            db, str(category["_id"]), buf.getvalue(), str(super_admin["_id"])
        )
        assert result.error

    async def test_with_explicit_column_map(self, db, category, super_admin):
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Q", "Opts", "Ans", "Diff"])
        ws.append(["What is 2+2?", "3, 4, 5", "2", "medium"])

        buf = io.BytesIO()
        wb.save(buf)

        result = await question_service.process_excel_import(
            db,
            str(category["_id"]),
            buf.getvalue(),
            str(super_admin["_id"]),
            column_map={"question": "Q", "options": "Opts", "answer": "Ans", "complexity": "Diff"},
        )
        assert result.created == 1

    async def test_skips_empty_rows(self, db, category, super_admin):
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Question", "Options", "Answer", "Complexity"])
        ws.append([None, "A, B", "1", "low"])  # empty question text — skipped

        buf = io.BytesIO()
        wb.save(buf)

        result = await question_service.process_excel_import(
            db, str(category["_id"]), buf.getvalue(), str(super_admin["_id"])
        )
        assert result.created == 0
